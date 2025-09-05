import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

import os
import sqlite3
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from libraries.duplikacio_config import DuplikacioConfig

def create_excel_export(excel_filename=None, output_folder=None):
    try:
        os.environ['DUPLIKACIO_SILENT'] = '1'
        config = DuplikacioConfig()
        # print(f"[DEBUG] CWD: {os.getcwd()}")
        if output_folder is None:
            output_folder = config.get_output_folder()
        # print(f"[DEBUG] output_folder: {output_folder}")
        if excel_filename is None:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            prefix = config.get_excel_prefix()
            excel_filename = f"{prefix}_{timestamp}.xlsx"
        # print(f"[DEBUG] excel_filename: {excel_filename}")
        full_path = os.path.join(output_folder, excel_filename)
        # print(f"[DEBUG] Excel export teljes elérési út: {full_path}")
        os.makedirs(os.path.dirname(full_path), exist_ok=True)
        conn = sqlite3.connect(config.get_database_file())
        cursor = conn.cursor()
        # Ellenőrizzük, hogy létezik-e a redundancia tábla, és ha nem, akkor hozzuk létre a helyes szerkezettel
        cursor.execute("""
            SELECT name FROM sqlite_master WHERE type='table' AND name='redundancia'
        """)
        table_exists = cursor.fetchone()
        create_sql = '''
            CREATE TABLE redundancia (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                status TEXT DEFAULT 'Rendben',
                file_path TEXT NOT NULL,
                file_name TEXT NOT NULL,
                file_size INTEGER NOT NULL,
                record_date TEXT NOT NULL,
                record_time TEXT NOT NULL,
                max_ismetlesek_szama INTEGER DEFAULT 0,
                max_ismetelt_karakterszam INTEGER DEFAULT 0,
                overview TEXT DEFAULT ''
            )
        '''
        if not table_exists:
            print('CREATE TABLE redundancia végrehajtva:')
            print(create_sql)
            cursor.execute(create_sql)
        try:
            os.environ['DUPLIKACIO_SILENT'] = '1'
            config = DuplikacioConfig()
            # print(f"[DEBUG] CWD: {os.getcwd()}")
            if output_folder is None:
                output_folder = config.get_output_folder()
            # print(f"[DEBUG] output_folder: {output_folder}")
            if excel_filename is None:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                prefix = config.get_excel_prefix()
                excel_filename = f"{prefix}_{timestamp}.xlsx"
            # print(f"[DEBUG] excel_filename: {excel_filename}")
            full_path = os.path.join(output_folder, excel_filename)
            # print(f"[DEBUG] Excel export teljes elérési út: {full_path}")
            os.makedirs(os.path.dirname(full_path), exist_ok=True)
            conn = sqlite3.connect(config.get_database_file())
            cursor = conn.cursor()
            # Ellenőrizzük, hogy létezik-e a redundancia tábla, és ha nem, akkor hozzuk létre a helyes szerkezettel
            cursor.execute("""
                SELECT name FROM sqlite_master WHERE type='table' AND name='redundancia'
            """)
            table_exists = cursor.fetchone()
            create_sql = '''
                CREATE TABLE redundancia (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    status TEXT DEFAULT 'Rendben',
                    file_path TEXT NOT NULL,
                    file_name TEXT NOT NULL,
                    file_size INTEGER NOT NULL,
                    record_date TEXT NOT NULL,
                    record_time TEXT NOT NULL,
                    max_ismetlesek_szama INTEGER DEFAULT 0,
                    max_ismetelt_karakterszam INTEGER DEFAULT 0,
                    overview TEXT DEFAULT ''
                )
            '''
            if not table_exists:
                print('CREATE TABLE redundancia végrehajtva:')
                print(create_sql)
                cursor.execute(create_sql)
            # hashCodes tábla létrehozása
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS hashCodes (
                    hash_value TEXT(100) PRIMARY KEY,
                    file_name TEXT NOT NULL,
                    file_path TEXT NOT NULL,
                    created_date TEXT NOT NULL,
                    created_time TEXT NOT NULL,
                    used_by_nbr INTEGER DEFAULT 0,
                    line_content TEXT,
                    redundancia_id INTEGER,
                    FOREIGN KEY (redundancia_id) REFERENCES redundancia(id)
                )
            ''')
            # repeat tábla létrehozása
            cursor.execute('''
                CREATE TABLE IF NOT EXISTS repeat (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    file_name TEXT NOT NULL,
                    source_file_path TEXT,
                    source_file_name TEXT NOT NULL,
                    redundancia_id INTEGER,
                    block_id INTEGER NOT NULL,
                    line_length INTEGER NOT NULL,
                    sum_line_length INTEGER DEFAULT 0,
                    repeated_line TEXT NOT NULL,
                    created_date TEXT NOT NULL,
                    created_time TEXT NOT NULL,
                    FOREIGN KEY (redundancia_id) REFERENCES redundancia(id)
                )
            ''')
            cursor.execute('''
                SELECT id, status, file_path, file_name, file_size, record_date, record_time, 
                       max_ismetlesek_szama, max_ismetelt_karakterszam, overview 
                FROM redundancia 
                ORDER BY max_ismetelt_karakterszam DESC
            ''')
            data = cursor.fetchall()
            # Excel fájl létrehozása és mentése
            wb = Workbook()
            ws = wb.active
            ws.title = "Redundancia"
            # Fejléc
            headers = ["ID", "Státusz", "Fájl elérési út", "Fájlnév", "Méret", "Dátum", "Idő", "Max ismétlés", "Max ismételt karakterszám", "Áttekintés"]
            ws.append(headers)
            # Adatok
            for row in data:
                ws.append(row)
            wb.save(full_path)
            # print(f"[DEBUG] Excel export sikeresen elmentve: {full_path}")
        except Exception as e:
            print(f"[DEBUG] Hiba történt az excel export során: {e}")
            return None
        ws.append(headers)
        # Adatok
        for row in data:
            ws.append(row)
        wb.save(full_path)
    # print(f"[DEBUG] Excel export sikeresen elmentve: {full_path}")
    except Exception as e:
    # print(f"[DEBUG] Hiba történt az excel export során: {e}")
        return None

# --- Főprogram blokk a közvetlen futtatáshoz ---
if __name__ == "__main__":
    create_excel_export()
    # print('[DEBUG] excel_export_simple.py modul betöltve')
